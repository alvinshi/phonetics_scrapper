import requests
import re
from openpyxl import Workbook
from openpyxl import load_workbook


class Scrapper:
    SUCCESS = 200
    WORD_PLACEHOLDER = "{word}"
    OXFORD_URL_TEMPLATE = "https://en.oxforddictionaries.com/definition/{}".format(WORD_PLACEHOLDER)
    OXFORD_REGEX = re.compile('<span class="phoneticspelling">(.*?)</span>')
    LONGMAN_URL_TEMPLATE = "https://www.ldoceonline.com/dictionary/{}".format(WORD_PLACEHOLDER)
    LONGMAN_REGEX = re.compile('<span class="PRON">(.*?)</span>')
    INPUT_PATH = "input.xlsx"
    OUTPUT_PATH = "output.xlsx"

    @staticmethod
    def _retrieve_html(url, add_headers=False):
        headers = {'Host': 'www.ldoceonline.com',
                   'Connection': 'keep-alive',
                   'Cache-Control': 'max-age=0',
                   'Upgrade-Insecure-Requests': '1',
                   'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/67.0.3396.99 Safari/537.36',
                   'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
                   'Accept-Encoding': 'gzip, deflate, br',
                   'Accept-Language': 'en-US,en;q=0.9',
                   'Cookie': '_ga=GA1.3.31325880.1530860210; _gid=GA1.3.70998704.1530860210; _gat=1; __qca=P0-418125505-1530860209666; __gads=ID=3934ca0d8c0095e6:T=1530860209:S=ALNI_MZJ-NqmC-EibvJLX6nemhGl3t7Q6g'
                   }
        if add_headers:
            response = requests.get(url, headers = headers)
        else:
            response = requests.get(url)
        return response.status_code, response.content

    @staticmethod
    def _get_cell_index(row, col):
        return col.upper() + str(row)

    def _get_words(self):
        wb = load_workbook(filename= self.INPUT_PATH)
        sheet = wb['Sheet1']
        row = 1
        words = []
        end_reached = False
        while not end_reached:
            cell = self._get_cell_index(row, "A")
            word = sheet[cell].value
            if word and len(word) > 0:
                words.append(word.lower())
                row += 1
            else:
                end_reached = True
        return words

    def _get_oxford_phonetics(self, word):
        url = self.OXFORD_URL_TEMPLATE.replace(self.WORD_PLACEHOLDER, word)
        _, content = self._retrieve_html(url)
        m = self.OXFORD_REGEX.search(content)
        if m:
            return m.group(1)
        else:
            return ""

    def _extract_longman_words(self, start, content):
        started = False
        tag_count = 0
        index = start
        should_record = False
        phonetics = ""
        while not (tag_count == 0 and started):
            started = True
            char = content[index]
            if char == '<':
                should_record = False
                if content[index + 1] == "s":
                    tag_count += 1
                else:
                    tag_count -= 1
            elif content[index - 1] == '>':
                should_record = True
            if should_record:
                phonetics += char
            index += 1
        phonetics = phonetics.replace('"',"")
        return "/" + phonetics + "/"

    def _get_longman_phonetics(self, word):
        url = self.LONGMAN_URL_TEMPLATE.replace(self.WORD_PLACEHOLDER, word)
        _, content = self._retrieve_html(url, True)
        m = self.LONGMAN_REGEX.search(content)
        if m:
            return self._extract_longman_words(m.start(), content)
        else:
            return ""

    def run(self):
        wb = Workbook()
        worksheet = wb.active
        words = self._get_words()
        errors = 0
        for index, word in enumerate(words):
            oxford_phonetics = self._get_oxford_phonetics(word)
            longman_phonetics = self._get_longman_phonetics(word)
            if oxford_phonetics == "" or longman_phonetics == "":
                errors += 1
            worksheet.cell(row=index+1, column=1, value=words[index])
            worksheet.cell(row=index+1, column=2, value=oxford_phonetics)
            worksheet.cell(row=index+1, column=3, value=longman_phonetics)
            if index % 10 == 0 and index > 0:
                print("{} done".format(index + 1))
                wb.save(self.OUTPUT_PATH)
        wb.save(self.OUTPUT_PATH)
        print("{} words have been matched with their phonetics".format(len(words) - errors))
        print("{} words failed".format(errors))
        wb.close()


def main():
    Scrapper().run()


if __name__ == "__main__":
    main()